import os
from openpyxl import Workbook, load_workbook
from typing import List

class SaveFile:
    _instance = None
    
    @staticmethod
    def get_save_file():
        if SaveFile._instance is None:
            SaveFile._instance = SaveFile()
        return SaveFile._instance
    
    @staticmethod
    def create_result_file(path_file: str):
        workbook = Workbook()
        workbook.save(path_file)

    @staticmethod
    def save_results_to_excel(run: int, path_file_end: str, heuristic: str, execution_times: List[float], unassigned_customers: List[int], id_dep_without_cust: List[int]):
        
        if os.path.exists(path_file_end):
            workbook = load_workbook(path_file_end)
        else:
            workbook = Workbook()

        sheet_name = f"Resultados_{heuristic.value.split('.')[-1]}"
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["RUN", "TIME(ms)", "TIME(s)", "UNA_CTS", "DPTS_WO"])
            
        for i in range(run):
            sheet.append([i + 1, execution_times[i], execution_times[i] * 1000.0, unassigned_customers[i], id_dep_without_cust[i] if i < len(id_dep_without_cust) else 0])

        max_time = max(execution_times, default=0)
        min_time = min(execution_times, default=0)
        avg_time = sum(execution_times) / len(execution_times) if execution_times else 0
        avg_time_seconds = avg_time / 1000.0

        max_unassigned = max(unassigned_customers, default=0)
        min_unassigned = min(unassigned_customers, default=0)
        avg_unassigned = sum(unassigned_customers) / len(unassigned_customers) if unassigned_customers else 0

        max_dep_without_cust = max(id_dep_without_cust, default=0) if id_dep_without_cust else 0
        min_dep_without_cust = min(id_dep_without_cust, default=0) if id_dep_without_cust else 0
        avg_dep_without_cust = sum(id_dep_without_cust) / len(id_dep_without_cust) if id_dep_without_cust else 0

        sheet.append([])
        sheet.append(["MAX:", max_time, max_time / 1000.0, max_unassigned, max_dep_without_cust])
        sheet.append(["MIN:", min_time, min_time / 1000.0, min_unassigned, min_dep_without_cust])
        sheet.append(["AVG:", avg_time, avg_time_seconds, avg_unassigned, avg_dep_without_cust])

        workbook.save(path_file_end)
        print(f"Resultados guardados en: {path_file_end}")